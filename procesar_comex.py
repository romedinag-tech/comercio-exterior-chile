# -*- coding: utf-8 -*-
"""
procesar_comex.py
Procesa los microdatos oficiales del Servicio Nacional de Aduanas (DUS exportaciones /
DIN importaciones, anio 2025) y construye el perfil de comercio exterior de cada punto
de transferencia de carga de Chile (puertos maritimos, aeropuertos, avanzadas fronterizas,
ferrocarriles / puertos secos).

Salidas:
  data/puntos.json            -> un registro por punto (perfil completo)
  data/puntos_tidy.csv        -> equivalente tabular
  data/paises_comercio.json   -> agregado pais-socio (nacional)
  data/meta.json              -> cuadre de control, fechas, fuentes, puntos sin coordenada
  data_bundle.js              -> bundle embebible para el HTML (puntos + paises + geojson + meta)

Regla de oro: NO se inventan datos. Todo monto/peso sale del microdato. Las coordenadas son
referencia geografica (tabla del spec). Los puntos sin coordenada se listan, no se inventan.

Fuente: Servicio Nacional de Aduanas, microdatos DUS/DIN 2025; clasificador 2022 v2.0 y
tablas de codigos del Servicio. Elaboracion propia.
"""
import csv, json, unicodedata, sys, os, re, hashlib
from collections import defaultdict, Counter
import openpyxl
import warnings
warnings.filterwarnings("ignore")

csv.field_size_limit(10_000_000)

# ---------------------------------------------------------------- rutas
HERE   = os.path.dirname(os.path.abspath(__file__))
STUDY  = os.path.dirname(HERE)
COMEX  = os.path.join(STUDY, "Antecedentes", "Estadisticas COMEX")
EXPO   = os.path.join(COMEX, "Exportaciones", "salidas2025", "Salidas2025.csv")
IMPO   = os.path.join(COMEX, "Importaciones", "Por lugar e ingreso", "ingresos_2025", "ingresos_2025.csv")
TABLAS = os.path.join(COMEX, "Exportaciones", "tablas_de_codigos.xlsx")
CLASIF = os.path.join(COMEX, "Exportaciones", "clasificador2022_v2_0.xlsx")
DATA   = os.path.join(HERE, "data");   os.makedirs(DATA, exist_ok=True)
ASSETS = os.path.join(HERE, "assets")

YEAR = 2025
CHILE_PAIS = 997

EXPORT_OPS = {200,201,202,203,204,205,206,207,210,211,212,213,216}
IMPORT_OPS = {101,102,103,104,105,113,115,116,119,120,121,122,123,129,130,134,142,151,152,165,171,179,180}

# aduana de tramitacion -> codigo de region (proxy de destino de importacion)
ADUANA_REGION = {
    3:15, 7:1, 10:2, 14:2, 17:3, 25:4, 33:5, 34:5, 39:5,
    48:13, 55:8, 67:10, 69:10, 83:11, 90:11, 92:12,
}

# Tabla de coordenadas = referencia geografica (del spec). Match por substring normalizado.
COORDS = {
 "ARICA": (-18.476,-70.323), "IQUIQUE": (-20.213,-70.152), "PATILLOS": (-20.74,-70.20),
 "TOCOPILLA": (-22.092,-70.197), "MEJILLONES": (-23.10,-70.45), "ANGAMOS": (-23.06,-70.43),
 "ANTOFAGASTA": (-23.65,-70.40), "CALDERA": (-27.067,-70.823), "BARQUITO": (-26.36,-70.66),
 "CHANARAL": (-26.35,-70.62), "HUASCO": (-28.46,-71.22), "GUACOLDA": (-28.46,-71.22),
 "COQUIMBO": (-29.95,-71.34), "GUAYACAN": (-29.97,-71.36), "VENTANAS": (-32.75,-71.48),
 "QUINTERO": (-32.78,-71.53), "VALPARAISO": (-33.04,-71.63), "SAN ANTONIO": (-33.59,-71.61),
 "LIRQUEN": (-36.71,-72.98), "PENCO": (-36.74,-72.99), "TALCAHUANO": (-36.72,-73.11),
 "SAN VICENTE": (-36.73,-73.13), "HUACHIPATO": (-36.74,-73.13), "CORONEL": (-37.03,-73.15),
 "LOTA": (-37.09,-73.16), "PUERTO MONTT": (-41.48,-72.94), "CALBUCO": (-41.77,-73.13),
 "CHACABUCO": (-45.46,-72.82), "PUERTO AYSEN": (-45.40,-72.69), "PUNTA ARENAS": (-53.16,-70.92),
 "NATALES": (-51.73,-72.51), "PUERTO WILLIAMS": (-54.93,-67.61), "CABO NEGRO": (-52.95,-70.80),
 "GREGORIO": (-52.62,-70.20), "TERMINAL GRANELES DEL NORTE": (-23.06,-70.43),
 "MERINO BENITEZ": (-33.39,-70.79), "ARTURO MERINO": (-33.39,-70.79),
 "CHACALLUTA": (-18.35,-70.34), "DIEGO ARACENA": (-20.54,-70.18), "CERRO MORENO": (-23.45,-70.44),
 "CARRIEL SUR": (-36.77,-73.06), "EL TEPUAL": (-41.44,-73.09), "IBANEZ DEL CAMPO": (-53.00,-70.85),
 "VISVIRI": (-17.60,-69.48), "CHUNGARA": (-18.22,-69.20), "COLCHANE": (-19.28,-68.64),
 "OLLAGUE": (-21.23,-68.25), "JAMA": (-23.23,-67.05), "SAN PEDRO DE ATACAMA": (-22.91,-67.99),
 "HITO CAJON": (-22.91,-67.77), "CONCORDIA": (-18.35,-70.32), "AGUA NEGRA": (-30.18,-69.82),
 "LOS LIBERTADORES": (-32.84,-70.08), "CRISTO REDENTOR": (-32.84,-70.08),
 "PEHUENCHE": (-35.98,-70.40), "PINO HACHADO": (-38.66,-70.89), "LIUCURA": (-38.66,-70.89),
 "CARDENAL SAMORE": (-40.72,-71.94), "FUTALEUFU": (-43.18,-71.85), "HUEMULES": (-45.62,-71.55),
 "COYHAIQUE ALTO": (-45.49,-71.30), "INTEGRACION AUSTRAL": (-52.16,-69.49),
 "MONTE AYMOND": (-52.16,-69.49), "DOROTEA": (-51.66,-72.36), "SAN SEBASTIAN": (-53.32,-68.62),
}

# Coordenadas de referencia ANADIDAS (ubicaciones geograficas publicas y conocidas, no dato COMEX).
# Extienden la tabla del spec para puntos relevantes que quedaban sin geolocalizar.
# Se marcan como "referencia anadida" en meta para trazabilidad.
COORDS_EXTRA = {
 "CALETA COLOSO": (-23.758,-70.470),   # terminal minero al sur de Antofagasta (Escondida)
 "COLOSO": (-23.758,-70.470),
 "PATACHE": (-20.810,-70.200),         # Punta Patache, al sur de Iquique
 "LOS VILOS": (-31.911,-71.510),
 "TALTAL": (-25.411,-70.484),
 "CORRAL": (-39.887,-73.430),
 "CONSTITUCION": (-35.333,-72.411),
 "JUAN FERNANDEZ": (-33.638,-78.830),
 "ISLA DE PASCUA": (-27.150,-109.433),
 "SAN FRANCISCO": (-26.900,-68.270),   # Paso San Francisco (Atacama)
 "MAHUIL MALAL": (-39.570,-71.500),    # Paso Mamuil Malal / Tromen
 "PALENA": (-43.620,-71.720),          # Paso Palena - Carrenleufu
 # --- completadas con fuente Servicio Nacional de Aduanas (ubicacion geografica de pasos) ---
 "ABRA DE NAPA": (-20.500,-68.583),    # Aduana: 20 30'S 68 35'W (altiplano, frontera con Bolivia)
 "LAGO VERDE": (-44.250,-71.800),      # Aduana: 44 15'S 71 48'W (Aysen)
 "PANGUIPULLI": (-39.643,-72.334),     # localidad (no figura en tabla de pasos de Aduana)
 "TERRITORIO ANTARTICO": (-62.196,-58.962),  # Base Pdte. Frei / Villa Las Estrellas, Isla Rey Jorge (punto representativo)
 "BAKER": (-47.150,-72.550),           # Aduana lat 47 09'S; long oficial 79 51'W es erronea (cae en el oceano); ubicado en zona Rio Baker / Cochrane (Aysen)
}
COORDS.update(COORDS_EXTRA)
COORDS_EXTRA_KEYS = set(COORDS_EXTRA)

# nombre Aduanas (normalizado, sin acentos, mayusculas) -> ISO3 del GeoJSON mundial
PAIS_ISO3 = {
 "CHINA":"CHN","ESTADOS UNIDOS DE AMERICA":"USA","JAPON":"JPN","BRASIL":"BRA",
 "COREA DEL SUR":"KOR","INDIA":"IND","HOLANDA":"NLD","ESPANA":"ESP","MEXICO":"MEX",
 "SUIZA":"CHE","PERU":"PER","ALEMANIA":"DEU","FRANCIA":"FRA","CANADA":"CAN",
 "ARGENTINA":"ARG","COLOMBIA":"COL","TAIWAN (FORMOSA)":"TWN","ITALIA":"ITA",
 "REINO UNIDO":"GBR","THAILANDIA":"THA","ECUADOR":"ECU","BELGICA":"BEL","RUSIA":"RUS",
 "SUECIA":"SWE","COSTA RICA":"CRI","VIETNAM":"VNM","BOLIVIA":"BOL","NORUEGA":"NOR",
 "AUSTRALIA":"AUS","POLONIA":"POL","FINLANDIA":"FIN","GUATEMALA":"GTM","BULGARIA":"BGR",
 "URUGUAY":"URY","MALASIA":"MYS","REPUBLICA DOMINICANA":"DOM","PARAGUAY":"PRY",
 "EMIRATOS ARABES UNIDOS":"ARE","BAHREIN":"BHR","DINAMARCA":"DNK","COSTA DE MARFIL":"CIV",
 "PUERTO RICO":"PRI","TURQUIA":"TUR","PANAMA":"PAN","NIGERIA":"NGA","ISRAEL":"ISR",
 "VENEZUELA":"VEN","FILIPINAS":"PHL","INDONESIA":"IDN","SINGAPUR":"SGP",
 "ARABIA SAUDITA":"SAU","SUDAFRICA":"ZAF","EGIPTO":"EGY","GRECIA":"GRC","PORTUGAL":"PRT",
 "IRLANDA":"IRL","AUSTRIA":"AUT","REPUBLICA CHECA":"CZE","HUNGRIA":"HUN","RUMANIA":"ROU",
 "UCRANIA":"UKR","HONDURAS":"HND","EL SALVADOR":"SLV","NICARAGUA":"NIC","CUBA":"CUB",
 "JAMAICA":"JAM","TRINIDAD Y TOBAGO":"TTO","NUEVA ZELANDIA":"NZL","NUEVA ZELANDA":"NZL",
 "PAKISTAN":"PAK","BANGLADESH":"BGD","SRI LANKA":"LKA","MARRUECOS":"MAR","ARGELIA":"DZA",
 "TUNEZ":"TUN","KENIA":"KEN","GHANA":"GHA","ANGOLA":"AGO","TANZANIA":"TZA",
 "ESLOVAQUIA":"SVK","ESLOVENIA":"SVN","CROACIA":"HRV","SERBIA":"SRB","LITUANIA":"LTU",
 "LETONIA":"LVA","ESTONIA":"EST","ISLANDIA":"ISL","LUXEMBURGO":"LUX","CHIPRE":"CYP",
 "QATAR":"QAT","KUWAIT":"KWT","OMAN":"OMN","JORDANIA":"JOR","LIBANO":"LBN","IRAN":"IRN",
 "IRAK":"IRQ","MYANMAR (BIRMANIA)":"MMR","CAMBOYA":"KHM","KAZAJSTAN":"KAZ",
 "HONG KONG":"HKG","MACAO":"HKG","BELICE":"BLZ","GUYANA":"GUY","SURINAM":"SUR",
 "MOZAMBIQUE":"MOZ","NAMIBIA":"NAM","ZAMBIA":"ZMB","ZIMBABWE":"ZWE","SENEGAL":"SEN",
 "CAMERUN":"CMR","GABON":"GAB","ETIOPIA":"ETH","MADAGASCAR":"MDG","MAURICIO":"MUS",
}

# ---------------------------------------------------------------- helpers
def norm(s):
    if s is None: return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s

def fnum(x):
    if x is None: return 0.0
    x = str(x).strip().replace(",", ".")
    if x == "": return 0.0
    try: return float(x)
    except ValueError: return 0.0

def fint(x):
    try: return int(str(x).strip())
    except (ValueError, TypeError): return None

def tipo_grupo(tipo):
    t = norm(tipo)
    if "MARITIMO" in t: return "maritimo"
    if "AEROPUERTO" in t or "AEREO" in t: return "aereo"
    return "terrestre"   # avanzada fronteriza, ferrocarril, puerto seco

def match_coord(nombre):
    n = norm(nombre)
    for k,(la,lo) in COORDS.items():
        if k in n or n in k:
            return la, lo, k
    # match por palabra significativa
    for k,(la,lo) in COORDS.items():
        for w in k.split():
            if len(w) >= 5 and w in n:
                return la, lo, k
    return None

# ---------------------------------------------------------------- diccionarios
def load_dicts():
    wb = openpyxl.load_workbook(TABLAS, read_only=True, data_only=True)
    def sheet(pref):
        for n in wb.sheetnames:
            if norm(n).startswith(norm(pref)): return wb[n]
        raise KeyError(pref)

    puertos = {}
    for r in sheet("Puertos").iter_rows(min_row=6, values_only=True):
        cod = fint(r[1])
        if cod is None: continue
        tipo = r[3]
        # correccion de fuente: el SNA clasifica "AEROPUERTO CARRIEL SUR 945" como
        # "Puerto maritimo"; si el nombre dice aeropuerto, se trata como tal
        if isinstance(r[2], str) and "AEROP" in norm(r[2]) and tipo and "MARITIMO" in norm(tipo):
            tipo = "Aeropuerto"
        puertos[cod] = dict(nombre=r[2], tipo=tipo, cod_pais=fint(r[4]),
                            pais=r[5], zona=r[6])
    paises = {}
    for r in sheet("Pais").iter_rows(min_row=6, values_only=True):
        cod = fint(r[1])
        if cod is None: continue
        paises[cod] = dict(nombre=r[2], continente=r[3])
    regiones = {}
    for r in sheet("Regiones").iter_rows(min_row=5, values_only=True):
        cod = fint(r[1])
        if cod is None: continue
        regiones[cod] = r[2]
    vias = {}
    for r in sheet("Vias").iter_rows(min_row=5, values_only=True):
        cod = fint(r[1])
        if cod is None: continue
        vias[cod] = r[2]
    tcarga = {}
    for r in sheet("Tipos de Carga").iter_rows(min_row=7, values_only=True):
        c = r[1]
        if isinstance(c, str) and len(c.strip()) == 1:
            tcarga[c.strip().upper()] = r[2]
    clausulas = {}
    for r in sheet("Clausulas").iter_rows(min_row=5, values_only=True):
        cod = fint(r[1])
        if cod is None: continue
        sig = r[3] if isinstance(r[3], str) and len(r[3].strip()) <= 5 else None
        clausulas[cod] = (sig.strip() if sig else r[2])
    wb.close()

    # clasificador: arancel(8) -> niveles
    wc = openpyxl.load_workbook(CLASIF, read_only=True, data_only=True)
    ws = wc[wc.sheetnames[0]]
    clasif = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ar = r[1]
        if ar is None: continue
        key = str(ar).strip().zfill(8)
        clasif[key] = dict(cap=str(r[0]).strip() if r[0] is not None else "",
                           n1_imp=r[3], n2_imp=r[4], n1_exp=r[6], n2_exp=r[7])
    wc.close()
    return puertos, paises, regiones, clasif, vias, tcarga, clausulas

# ---------------------------------------------------------------- series anuales por puerto (agregado oficial)
ANUAL_YEARS = list(range(2012, 2026))   # 2012..2025

def nkey(s):
    s = norm(s)
    return "".join(ch for ch in s if ch.isalnum())

# nombre del punto (tabla Puertos) -> nombre en los archivos agregados (cuando difieren)
# cada punto puede tener VARIOS nombres segun el archivo (2002-2021 vs 2022-2025)
ANUAL_ALIASES = {
    "AEROP. A.M. BENITEZ": ["Aeropuerto Arturo Merino Benitez", "AEROP. A. MERINO BENITEZ"],
    "AEROP. CERRO MORENO": ["Aeropuerto Cerro Moreno", "AEROP. CERRO MORENO"],
    "AEROPUERTO CARRIEL SUR 945": ["Aeropuerto Carriel Sur", "AEROP. CARRIEL SUR"],
    "AEROP. C.I. DEL CAMPO": ["Aeropuerto Carlos Ibanez del Campo", "AEROP. C. IBANEZ DEL CAMPO"],
    "AEROP. CHACALLUTA": ["Aeropuerto Chacalluta"],
    "AEROP. DIEGO ARACENA": ["Aeropuerto Diego Aracena"],
    "AEROP. EL TEPUAL": ["Aeropuerto El Tepual"],
    "CAP. HUACHIPATO": ["Muelle Huachipato"],
    "LOS LIBERTADORES": ["Cristo Redentor (Los Libertadores)"],
    "MONTE AYMOND": ["Integracion Austral (Monte Aymond)"],
    "PASO JAMA": ["Jama"],
    "CHACALLUTA": ["Concordia (Chacalluta)"],
}

def read_pto_year(path):
    """Lee un archivo agregado 'por puerto x anio'. Autodetecta fila de encabezado
    y columna de nombre. Devuelve {n2(nombre): {anio: valor}} y la fila 'Total'."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = None; ycols = []
    for i, r in enumerate(rows):
        yc = [(j, int(str(c))) for j, c in enumerate(r)
              if c is not None and str(c).strip().isdigit() and 1990 < int(str(c)) < 2030]
        if len(yc) >= 3:
            hdr = i; ycols = yc; break
    if hdr is None:
        return {}
    namecol = min(j for j, _ in ycols) - 1
    out = {}
    for r in rows[hdr+1:]:
        if namecol >= len(r): continue
        name = r[namecol]
        if name is None or nkey(name) in ("", "IRALISTADODETABLAS"): continue
        if isinstance(name, str) and ("Fuente:" in name or "Declaraciones" in name): continue
        vals = {}
        for j, y in ycols:
            v = r[j] if j < len(r) else None
            vals[y] = fnum(v)
        out[nkey(name)] = vals
    return out

def merge_year(*dicts):
    m = {}
    for d in dicts:
        for k, vals in d.items():
            m.setdefault(k, {}).update(vals)
    return m

def load_anual():
    ex_fob = merge_year(
        read_pto_year(os.path.join(COMEX,"Exportaciones","expo_pto_monto_2002_2021.xlsx")),
        read_pto_year(os.path.join(COMEX,"Exportaciones","expo_pto_monto_2022_2025.xlsx")))
    im_cif = merge_year(
        read_pto_year(os.path.join(COMEX,"Importaciones","Por lugar e ingreso","impo_pto_monto_2002_2021.xlsx")),
        read_pto_year(os.path.join(COMEX,"impo_pto_monto_2022_2025.xlsx")))
    ex_kg = merge_year(
        read_pto_year(os.path.join(COMEX,"Exportaciones","expo_pto_peso_2002_2021.xlsx")),
        read_pto_year(os.path.join(COMEX,"Exportaciones","expo_pto_peso_2022_2025.xlsx")))
    im_kg = merge_year(
        read_pto_year(os.path.join(COMEX,"Importaciones","Por lugar e ingreso","impo_pto_peso_2002_2021.xlsx")),
        read_pto_year(os.path.join(COMEX,"Importaciones","Por lugar e ingreso","impo_pto_peso_2022_2025.xlsx")))
    return dict(ex_fob=ex_fob, im_cif=im_cif, ex_kg=ex_kg, im_kg=im_kg)

def anual_for(nombre, ANUAL):
    # busca bajo el nombre original Y el alias: los archivos 2002-2021 y 2022-2025
    # usan nomenclaturas distintas (p.ej. "AEROP.A.M.BENITEZ" vs "Aeropuerto Arturo Merino Benitez")
    als = ANUAL_ALIASES.get(norm(nombre), [])
    keys = {nkey(nombre)} | {nkey(a) for a in als}
    def serie(src):
        m = {}
        for k in keys:
            for y, v in ANUAL[src].get(k, {}).items():
                m[y] = max(m.get(y, 0.0), v)   # rangos de anios disjuntos; max evita doble conteo
        return [round(m.get(y, 0.0), 1) for y in ANUAL_YEARS]
    e = serie("ex_fob"); i = serie("im_cif")
    if not any(e) and not any(i):
        return None
    return dict(years=ANUAL_YEARS, exp_fob=e, imp_cif=i,
                exp_kg=serie("ex_kg"), imp_kg=serie("im_kg"))

print("Cargando diccionarios...", flush=True)
PUERTOS, PAISES, REGIONES, CLASIF, VIAS, TCARGA, CLAUS = load_dicts()
print("Cargando series anuales por puerto (2012-2025)...", flush=True)
ANUAL = load_anual()
CHILE_PORTS = {c for c,p in PUERTOS.items() if p["cod_pais"] == CHILE_PAIS}
print(f"  puertos CL={len(CHILE_PORTS)}  paises={len(PAISES)}  regiones={len(REGIONES)}  aranceles={len(CLASIF)}", flush=True)

def clasif_get(item, field):
    key = str(item).strip().zfill(8)
    rec = CLASIF.get(key)
    if rec is None:
        # intenta por capitulo (2 primeros digitos) -> no; devuelve sin clasificar
        return None
    return rec.get(field)

# ---------------------------------------------------------------- acumuladores
def new_port():
    return dict(fob=0.0, peso_exp=0.0, cif=0.0, peso_imp=0.0,
                dest=Counter(), orig=Counter(),
                texp=Counter(), timp=Counter(),
                pexp=Counter(), pimp=Counter(),
                pexp_kg=Counter(), pimp_kg=Counter(),
                reg_exp=Counter(), reg_imp=Counter(),
                via_exp=Counter(), via_imp=Counter(),
                tc_exp=Counter(), tc_imp=Counter(),
                cl_exp=Counter(), cl_imp=Counter(),
                advalorem=0.0,
                mes_exp=[0.0]*12, mes_imp=[0.0]*12)

ports = defaultdict(new_port)
# perfil por REGION (export: COD_REGION_ORIGEN real; import: region de la aduana de nacionalizacion)
regs = defaultdict(new_port)
reg_ptos_exp = defaultdict(Counter)   # region -> punto de salida -> fob
reg_ptos_imp = defaultdict(Counter)   # region -> punto de entrada -> cif
nat = dict(fob=0.0, cif=0.0, peso_exp=0.0, peso_imp=0.0,
           dest=Counter(), orig=Counter(),
           mes_exp=[0.0]*12, mes_imp=[0.0]*12,
           reg_exp=Counter(), reg_imp=Counter(),
           pexp=Counter(), pimp=Counter(), pexp_kg=Counter(), pimp_kg=Counter(),
           via_exp=Counter(), via_imp=Counter(), tc_exp=Counter(), tc_imp=Counter(),
           cl_exp=Counter(), cl_imp=Counter(), advalorem=0.0)
nat_fob_all = 0.0   # exportacion nacional total (todas las ops export, cualquier puerto)
nat_cif_all = 0.0

# ---------------------------------------------------------------- EXPORTACIONES
print("Procesando exportaciones (Salidas2025.csv)...", flush=True)
n=0; kept=0
with open(EXPO, encoding="latin-1", newline="") as f:
    rd = csv.reader(f, delimiter=";")
    next(rd, None)
    for row in rd:
        n += 1
        if len(row) < 17: continue
        op = fint(row[3])
        if op not in EXPORT_OPS: continue
        fob = fnum(row[14]); peso = fnum(row[16])
        nat_fob_all += fob
        port = fint(row[6])
        if port not in CHILE_PORTS:   # solo puntos chilenos para el mapa
            continue
        kept += 1
        pais = fint(row[8]); region = fint(row[4]); via = fint(row[5])
        item = row[13]; mes = fint(row[1])
        clau = fint(row[11]); tcar = (row[12] or "").strip().upper()
        d = ports[port]
        d["fob"] += fob; d["peso_exp"] += peso
        if mes and 1 <= mes <= 12: d["mes_exp"][mes-1] += fob
        if pais is not None: d["dest"][pais] += fob
        n1 = clasif_get(item, "n1_exp"); n2 = clasif_get(item, "n2_exp")
        if n1: d["texp"][n1] += fob
        if n2: d["pexp"][n2] += fob; d["pexp_kg"][n2] += peso
        if region is not None: d["reg_exp"][region] += fob
        if via is not None: d["via_exp"][via] += fob
        if tcar: d["tc_exp"][tcar] += fob
        if clau is not None: d["cl_exp"][clau] += fob
        nat["fob"] += fob; nat["peso_exp"] += peso
        if pais is not None: nat["dest"][pais] += fob
        if mes and 1 <= mes <= 12: nat["mes_exp"][mes-1] += fob
        if region is not None: nat["reg_exp"][region] += fob
        if n2: nat["pexp"][n2] += fob; nat["pexp_kg"][n2] += peso
        if via is not None: nat["via_exp"][via] += fob
        if tcar: nat["tc_exp"][tcar] += fob
        if clau is not None: nat["cl_exp"][clau] += fob
        if region is not None:
            rg = regs[region]
            rg["fob"] += fob; rg["peso_exp"] += peso
            if mes and 1 <= mes <= 12: rg["mes_exp"][mes-1] += fob
            if pais is not None: rg["dest"][pais] += fob
            if n1: rg["texp"][n1] += fob
            if n2: rg["pexp"][n2] += fob; rg["pexp_kg"][n2] += peso
            if via is not None: rg["via_exp"][via] += fob
            if tcar: rg["tc_exp"][tcar] += fob
            if clau is not None: rg["cl_exp"][clau] += fob
            reg_ptos_exp[region][port] += fob
print(f"  filas={n}  export-ops puntos CL={kept}  FOB nacional(all)={nat_fob_all:,.0f}", flush=True)

# ---------------------------------------------------------------- IMPORTACIONES
print("Procesando importaciones (ingresos_2025.csv)...", flush=True)
n=0; kept=0
with open(IMPO, encoding="latin-1", newline="") as f:
    rd = csv.reader(f, delimiter=";")
    next(rd, None)
    for row in rd:
        n += 1
        if len(row) < 17: continue
        op = fint(row[3])
        if op not in IMPORT_OPS: continue
        cif = fnum(row[12])
        nat_cif_all += cif
        port = fint(row[8])
        if port not in CHILE_PORTS:
            continue
        kept += 1
        pais = fint(row[4]); item = row[11]; via = fint(row[9]); aduana = fint(row[2])
        cant = fnum(row[15]); unidad = fint(row[16]); mes = fint(row[1])
        adval = fnum(row[13]); clau = fint(row[10]); tcar = (row[17] or "").strip().upper()
        peso = cant if unidad == 6 else 0.0
        d = ports[port]
        d["cif"] += cif; d["peso_imp"] += peso; d["advalorem"] += adval
        if mes and 1 <= mes <= 12: d["mes_imp"][mes-1] += cif
        if pais is not None: d["orig"][pais] += cif
        n1 = clasif_get(item, "n1_imp"); n2 = clasif_get(item, "n2_imp")
        if n1: d["timp"][n1] += cif
        if n2: d["pimp"][n2] += cif; d["pimp_kg"][n2] += peso
        reg = ADUANA_REGION.get(aduana)
        if reg is not None: d["reg_imp"][reg] += cif
        if via is not None: d["via_imp"][via] += cif
        if tcar: d["tc_imp"][tcar] += cif
        if clau is not None: d["cl_imp"][clau] += cif
        nat["cif"] += cif; nat["peso_imp"] += peso; nat["advalorem"] += adval
        if pais is not None: nat["orig"][pais] += cif
        if mes and 1 <= mes <= 12: nat["mes_imp"][mes-1] += cif
        if n2: nat["pimp"][n2] += cif; nat["pimp_kg"][n2] += peso
        if via is not None: nat["via_imp"][via] += cif
        if tcar: nat["tc_imp"][tcar] += cif
        if clau is not None: nat["cl_imp"][clau] += cif
        reg = ADUANA_REGION.get(aduana)
        if reg is not None:
            nat["reg_imp"][reg] += cif
            rg = regs[reg]
            rg["cif"] += cif; rg["peso_imp"] += peso; rg["advalorem"] += adval
            if mes and 1 <= mes <= 12: rg["mes_imp"][mes-1] += cif
            if pais is not None: rg["orig"][pais] += cif
            if n1: rg["timp"][n1] += cif
            if n2: rg["pimp"][n2] += cif; rg["pimp_kg"][n2] += peso
            if via is not None: rg["via_imp"][via] += cif
            if tcar: rg["tc_imp"][tcar] += cif
            if clau is not None: rg["cl_imp"][clau] += cif
            reg_ptos_imp[reg][port] += cif
        if n % 300000 == 0: print(f"    ...{n} filas", flush=True)
print(f"  filas={n}  import-ops puntos CL={kept}  CIF nacional(all)={nat_cif_all:,.0f}", flush=True)

# ---------------------------------------------------------------- armado de registros
def pais_nombre(c):
    r = PAISES.get(c); return r["nombre"] if r else f"PAIS_{c}"
def pais_iso3(c):
    r = PAISES.get(c)
    if not r: return None
    return PAIS_ISO3.get(norm(r["nombre"]))
def region_nombre(c):
    return REGIONES.get(c, f"REGION_{c}")

def top_pais(counter, total, k=8):
    out=[]
    for c,v in counter.most_common(k):
        out.append(dict(cod=c, nombre=pais_nombre(c), iso3=pais_iso3(c), valor=round(v,1),
                        pct=round(100*v/total,2) if total else 0))
    return out

def top_tipo(counter, total, k=8):
    out=[]
    for name,v in counter.most_common(k):
        out.append(dict(nombre=name, valor=round(v,1),
                        pct=round(100*v/total,2) if total else 0))
    return out

def top_prod(counter, total, k=12):
    out=[]; acum=0.0
    for name,v in counter.most_common(k):
        acum += v
        out.append(dict(nombre=name, valor=round(v,1),
                        pct=round(100*v/total,2) if total else 0,
                        pct_acum=round(100*acum/total,2) if total else 0))
    return out

def top_region(counter, total, k=16):
    out=[]
    for c,v in counter.most_common(k):
        out.append(dict(cod=c, nombre=region_nombre(c), valor=round(v,1),
                        pct=round(100*v/total,2) if total else 0))
    return out

def via_dominante(d):
    tot = d["via_exp"] + d["via_imp"]
    if not tot: return None
    return VIAS.get(tot.most_common(1)[0][0])

def top_named(counter, total, lookup, k=8):
    out=[]
    for cod,v in counter.most_common(k):
        out.append(dict(nombre=lookup(cod), valor=round(v,1),
                        pct=round(100*v/total,2) if total else 0))
    return out
def via_nombre(c):  return VIAS.get(c) or "Otra / sin info"
def tc_nombre(c):   return TCARGA.get(c) or "Otro / sin info"
def cl_nombre(c):   return CLAUS.get(c) or "Otra / sin info"

def hhi(counter):
    """Indice Herfindahl-Hirschman (0-10000). >2500 concentrado, <1500 diversificado."""
    tot = sum(counter.values())
    if tot <= 0: return 0
    return round(sum((v/tot)**2 for v in counter.values())*10000, 0)

def precio_prod(fobc, kgc, k=10):
    """Precio implicito US$/kg de los top productos por valor."""
    out=[]
    for name,fob in fobc.most_common(k):
        kg = kgc.get(name, 0.0)
        out.append(dict(nombre=name, valor=round(fob,1), kg=round(kg,1),
                        usd_kg=round(fob/kg, 2) if kg > 0 else None))
    return out

def estructura(d):
    """Bloque de estadisticas logisticas y de estructura por punto."""
    fob, cif = d["fob"], d["cif"]
    return dict(
        via_exp=top_named(d["via_exp"], fob, via_nombre, 6),
        via_imp=top_named(d["via_imp"], cif, via_nombre, 6),
        tcarga_exp=top_named(d["tc_exp"], fob, tc_nombre, 6),
        tcarga_imp=top_named(d["tc_imp"], cif, tc_nombre, 6),
        incoterm_exp=top_named(d["cl_exp"], fob, cl_nombre, 6),
        incoterm_imp=top_named(d["cl_imp"], cif, cl_nombre, 6),
        hhi=dict(destinos=hhi(d["dest"]), origenes=hhi(d["orig"]),
                 prod_exp=hhi(d["pexp"]), prod_imp=hhi(d["pimp"])),
        precio_exp=precio_prod(d["pexp"], d["pexp_kg"]),
        precio_imp=precio_prod(d["pimp"], d["pimp_kg"]),
        usd_kg_exp=round(fob/d["peso_exp"], 2) if d["peso_exp"] > 0 else None,
        usd_kg_imp=round(cif/d["peso_imp"], 2) if d["peso_imp"] > 0 else None,
        advalorem=round(d["advalorem"], 1),
        advalorem_pct=round(100*d["advalorem"]/cif, 2) if cif > 0 else 0,
    )

registros=[]; sin_coord=[]
for cod in sorted(ports):
    d = ports[cod]
    meta = PUERTOS[cod]
    total = d["fob"] + d["cif"]
    if total <= 0: continue
    coord = match_coord(meta["nombre"])
    if coord is None:
        lat, lon, coord_src = None, None, None
        sin_coord.append(dict(cod=cod, nombre=meta["nombre"], tipo=meta["tipo"],
                              total=round(total,0)))
    else:
        lat, lon, mkey = coord
        coord_src = "referencia_anadida" if mkey in COORDS_EXTRA_KEYS else "spec"
    reg = dict(
        cod=cod, nombre=meta["nombre"], tipo=meta["tipo"],
        grupo=tipo_grupo(meta["tipo"]), zona=meta["zona"],
        lat=lat, lon=lon, coord_src=coord_src,
        exp_fob=round(d["fob"],1), imp_cif=round(d["cif"],1),
        total=round(total,1),
        peso_exp=round(d["peso_exp"],1), peso_imp=round(d["peso_imp"],1),
        balance=round(d["cif"]-d["fob"],1),
        via_dom=via_dominante(d),
        top_destinos=top_pais(d["dest"], d["fob"]),
        top_origenes=top_pais(d["orig"], d["cif"]),
        top_tipo_exp=top_tipo(d["texp"], d["fob"]),
        top_tipo_imp=top_tipo(d["timp"], d["cif"]),
        top_prod_exp=top_prod(d["pexp"], d["fob"]),
        top_prod_imp=top_prod(d["pimp"], d["cif"]),
        infl_exp=top_region(d["reg_exp"], d["fob"]),   # regiones de origen de la carga exportada
        infl_imp=top_region(d["reg_imp"], d["cif"]),   # region de la aduana (proxy destino import)
        mes_exp=[round(x,1) for x in d["mes_exp"]],
        mes_imp=[round(x,1) for x in d["mes_imp"]],
        anual=anual_for(meta["nombre"], ANUAL),
        estructura=estructura(d),
    )
    registros.append(reg)

registros.sort(key=lambda r: r["total"], reverse=True)
sum_fob = sum(r["exp_fob"] for r in registros)
sum_cif = sum(r["imp_cif"] for r in registros)

# ---------------------------------------------------------------- perfil por REGION
port_coord = {r["cod"]: (r["lat"], r["lon"]) for r in registros}

# ubicacion geografica de cada punto -> region (point-in-polygon sobre el geojson regional)
_GEO_REG = json.load(open(os.path.join(ASSETS, "chile_regiones.geojson"), encoding="latin-1"))

def _pip(lon, lat, ring):
    inside = False; j = len(ring) - 1
    for i in range(len(ring)):
        xi, yi = ring[i][0], ring[i][1]; xj, yj = ring[j][0], ring[j][1]
        if (yi > lat) != (yj > lat) and lon < (xj - xi) * (lat - yi) / (yj - yi) + xi:
            inside = not inside
        j = i
    return inside

def _polys(g):
    return g["coordinates"] if g["type"] == "MultiPolygon" else [g["coordinates"]]

def punto_region(lat, lon):
    if lat is None: return None
    for f in _GEO_REG["features"]:
        for poly in _polys(f["geometry"]):
            if _pip(lon, lat, poly[0]):
                return f["properties"]["codregion"]
    # fallback: region con vertice mas cercano (puertos costeros que caen fuera del poligono)
    best, bd = None, 1e18
    for f in _GEO_REG["features"]:
        for poly in _polys(f["geometry"]):
            for x, y in poly[0][::4]:
                d = (x - lon) ** 2 + (y - lat) ** 2
                if d < bd: bd, best = d, f["properties"]["codregion"]
    return best

# puntos insulares / antarticos que el poligono continental no cubre
REGION_UBIC_OVERRIDE = {"ISLA DE PASCUA": 5, "JUAN FERNANDEZ": 5, "TERRITORIO ANTARTICO": 12}

for r in registros:
    ov = next((v for k, v in REGION_UBIC_OVERRIDE.items() if k in norm(r["nombre"])), None)
    r["region_ubicacion"] = ov if ov is not None else punto_region(r["lat"], r["lon"])

def top_puntos(counter, total, k=10):
    out=[]
    for pc, v in counter.most_common(k):
        info = PUERTOS.get(pc) or {}
        la, lo = port_coord.get(pc, (None, None))
        out.append(dict(cod=pc, nombre=info.get("nombre", f"PUNTO_{pc}"), lat=la, lon=lo,
                        valor=round(v,1), pct=round(100*v/total,2) if total else 0))
    return out

# serie anual regional = suma de las series de los terminales UBICADOS en la region
def anual_region(rc):
    pts = [r for r in registros if r.get("region_ubicacion") == rc and r.get("anual")]
    if not pts: return None
    def ssum(field):
        acc = [0.0] * len(ANUAL_YEARS)
        for r in pts:
            for i, v in enumerate(r["anual"][field]): acc[i] += v
        return [round(x, 1) for x in acc]
    return dict(years=ANUAL_YEARS, exp_fob=ssum("exp_fob"), imp_cif=ssum("imp_cif"),
                exp_kg=ssum("exp_kg"), imp_kg=ssum("imp_kg"),
                nota="Carga movilizada por los terminales ubicados en la región",
                n_terminales=len(pts))

regiones_perfil=[]
for rc in sorted(regs):
    d = regs[rc]
    rtotal = d["fob"] + d["cif"]
    if rtotal <= 0: continue
    regiones_perfil.append(dict(
        cod=rc, nombre=region_nombre(rc), tipo="Región", grupo="region",
        lat=None, lon=None, coord_src=None,
        exp_fob=round(d["fob"],1), imp_cif=round(d["cif"],1), total=round(rtotal,1),
        peso_exp=round(d["peso_exp"],1), peso_imp=round(d["peso_imp"],1),
        balance=round(d["cif"]-d["fob"],1), via_dom=via_dominante(d),
        top_destinos=top_pais(d["dest"], d["fob"]),
        top_origenes=top_pais(d["orig"], d["cif"]),
        top_tipo_exp=top_tipo(d["texp"], d["fob"]),
        top_tipo_imp=top_tipo(d["timp"], d["cif"]),
        top_prod_exp=top_prod(d["pexp"], d["fob"]),
        top_prod_imp=top_prod(d["pimp"], d["cif"]),
        top_puntos_exp=top_puntos(reg_ptos_exp[rc], d["fob"]),
        top_puntos_imp=top_puntos(reg_ptos_imp[rc], d["cif"]),
        mes_exp=[round(x,1) for x in d["mes_exp"]],
        mes_imp=[round(x,1) for x in d["mes_imp"]],
        anual=anual_region(rc), estructura=estructura(d),
    ))
print(f"  regiones con perfil: {len(regiones_perfil)}  con serie anual: {sum(1 for r in regiones_perfil if r['anual'])}", flush=True)

# ---------------------------------------------------------------- paises_comercio (nacional)
paises_comercio = dict(
    exp=[dict(cod=c, nombre=pais_nombre(c), iso3=pais_iso3(c),
              continente=(PAISES.get(c) or {}).get("continente"),
              fob=round(v,1), pct=round(100*v/nat["fob"],3) if nat["fob"] else 0)
         for c,v in nat["dest"].most_common()],
    imp=[dict(cod=c, nombre=pais_nombre(c), iso3=pais_iso3(c),
              continente=(PAISES.get(c) or {}).get("continente"),
              cif=round(v,1), pct=round(100*v/nat["cif"],3) if nat["cif"] else 0)
         for c,v in nat["orig"].most_common()],
)

# ---------------------------------------------------------------- meta / cuadre
meta = dict(
    anio=YEAR,
    generado="procesar_comex.py",
    n_puntos=len(registros),
    cuadre=dict(
        exp_nacional_total=round(nat_fob_all,0),
        suma_fob_puntos=round(sum_fob,0),
        exp_cobertura_pct=round(100*sum_fob/nat_fob_all,2) if nat_fob_all else 0,
        imp_nacional_total=round(nat_cif_all,0),
        suma_cif_puntos=round(sum_cif,0),
        imp_cobertura_pct=round(100*sum_cif/nat_cif_all,2) if nat_cif_all else 0,
    ),
    sin_coordenada=sin_coord,
    fuentes=[
        "Servicio Nacional de Aduanas - microdato DUS (exportaciones) 2025",
        "Servicio Nacional de Aduanas - microdato DIN (importaciones) 2025",
        "Servicio Nacional de Aduanas - tablas de codigos (Anexo 51)",
        "Servicio Nacional de Aduanas - clasificador 2022 v2.0",
        "Coordenadas: referencia geografica publica (tabla del estudio, extendida)",
    ],
    notas=[
        "Influencia de EXPORTACION (infl_exp): region de origen real de la carga "
        "(COD_REGION_ORIGEN del DUS). Responde 'de que region del pais proviene lo exportado'.",
        "Influencia de IMPORTACION (infl_imp): region de la ADUANA DE NACIONALIZACION "
        "(COD_ADUANA_TRAMITACION). Indica donde se nacionaliza la carga, NO la region de "
        "destino/consumo final, que el microdato DIN no registra.",
        "Peso de importacion: el DIN no trae peso bruto; se usa CANTIDAD_MERCANCIA solo cuando "
        "COD_UNIDAD_MEDIDA = 6 (kilogramos).",
        "Cobertura del cuadre <100%: corresponde a operaciones con puerto no chileno o "
        "codigo de puerto 0/desconocido en el microdato.",
        "Coordenadas completadas (Abra de Napa, Lago Verde, Baker) desde la tabla de ubicacion "
        "geografica de pasos del Servicio Nacional de Aduanas. La longitud oficial de BAKER "
        "(79 51'W) es erronea (cae en el oceano Pacifico); se conservo la latitud oficial "
        "(47 09'S) y se ubico en la zona del Rio Baker / Cochrane (Aysen). Panguipulli usa la "
        "localidad y Territorio Antartico la Base Frei (Isla Rey Jorge) como punto representativo.",
    ],
)

# ---------------------------------------------------------------- escritura
with open(os.path.join(DATA,"puntos.json"),"w",encoding="utf-8") as f:
    json.dump(registros, f, ensure_ascii=False, indent=1)
with open(os.path.join(DATA,"paises_comercio.json"),"w",encoding="utf-8") as f:
    json.dump(paises_comercio, f, ensure_ascii=False, indent=1)
with open(os.path.join(DATA,"meta.json"),"w",encoding="utf-8") as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

# csv tidy
with open(os.path.join(DATA,"puntos_tidy.csv"),"w",encoding="utf-8-sig",newline="") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["cod","nombre","tipo","grupo","zona","lat","lon",
                "exp_fob_usd","imp_cif_usd","total_usd","peso_exp_kg","peso_imp_kg",
                "balance_usd","top_destino","top_origen","top_tipo_exp","top_tipo_imp"])
    for r in registros:
        w.writerow([r["cod"],r["nombre"],r["tipo"],r["grupo"],r["zona"],r["lat"],r["lon"],
                    r["exp_fob"],r["imp_cif"],r["total"],r["peso_exp"],r["peso_imp"],r["balance"],
                    r["top_destinos"][0]["nombre"] if r["top_destinos"] else "",
                    r["top_origenes"][0]["nombre"] if r["top_origenes"] else "",
                    r["top_tipo_exp"][0]["nombre"] if r["top_tipo_exp"] else "",
                    r["top_tipo_imp"][0]["nombre"] if r["top_tipo_imp"] else ""])

# ---------------------------------------------------------------- geojson minificados + bundle
def round_geom(geom, nd=2):
    def rc(c):
        if isinstance(c[0], (int,float)):
            return [round(c[0],nd), round(c[1],nd)]
        return [rc(x) for x in c]
    geom["coordinates"] = rc(geom["coordinates"])
    return geom

# mundo: id ISO3 + name
world = json.load(open(os.path.join(ASSETS,"world_countries.geojson"),encoding="utf-8"))
for ft in world["features"]:
    ft["properties"] = {"name": ft["properties"].get("name")}
    round_geom(ft["geometry"], 2)

# chile regiones: codregion + nombre limpio (desde tabla SNA)
chreg = json.load(open(os.path.join(ASSETS,"chile_regiones.geojson"),encoding="latin-1"))
for ft in chreg["features"]:
    cr = ft["properties"].get("codregion")
    ft["properties"] = {"codregion": cr, "nombre": region_nombre(cr)}
    round_geom(ft["geometry"], 2)

# registro nacional (Total nacional) para la vista por defecto
nacional = dict(
    cod=0, nombre="TOTAL NACIONAL", tipo="Nacional", grupo="nacional", zona="Chile",
    lat=None, lon=None, coord_src=None,
    exp_fob=round(nat["fob"],1), imp_cif=round(nat["cif"],1),
    total=round(nat["fob"]+nat["cif"],1),
    peso_exp=round(nat["peso_exp"],1), peso_imp=round(nat["peso_imp"],1),
    balance=round(nat["cif"]-nat["fob"],1), via_dom=None,
    top_destinos=top_pais(nat["dest"], nat["fob"]),
    top_origenes=top_pais(nat["orig"], nat["cif"]),
    top_tipo_exp=[], top_tipo_imp=[], top_prod_exp=[], top_prod_imp=[],
    infl_exp=top_region(nat["reg_exp"], nat["fob"]),
    infl_imp=top_region(nat["reg_imp"], nat["cif"]),
    mes_exp=[round(x,1) for x in nat["mes_exp"]],
    mes_imp=[round(x,1) for x in nat["mes_imp"]],
    anual=None,
    estructura=estructura(nat),
)
# serie anual nacional = suma de los puntos (consistente con lo que se muestra)
def _sum_anual(field):
    acc=[0.0]*len(ANUAL_YEARS)
    for r in registros:
        a=r.get("anual")
        if not a: continue
        for k,v in enumerate(a[field]): acc[k]+=v
    return [round(x,1) for x in acc]
nacional["anual"]=dict(years=ANUAL_YEARS, exp_fob=_sum_anual("exp_fob"),
    imp_cif=_sum_anual("imp_cif"), exp_kg=_sum_anual("exp_kg"), imp_kg=_sum_anual("imp_kg"))

# ================= LOTE 2: estadisticas macro + trafico terrestre =================
print("Procesando estadisticas macro (lote 2)...", flush=True)
MYEARS = list(range(2012, 2026))
EX = os.path.join(COMEX, "Exportaciones")
IM = os.path.join(COMEX, "Importaciones")
PROD_E1=os.path.join(EX,"expo_prod_monto_2012_2021.xlsx"); PROD_E2=os.path.join(EX,"expo_prod_monto_2022_2025.xlsx")
PROD_I1=os.path.join(IM,"Importaciones por producto","impo_prod_monto_2012_2021.xlsx"); PROD_I2=os.path.join(IM,"Importaciones por producto","impo_prod_monto_2022_2025.xlsx")
PP_E1=os.path.join(EX,"expo_paisprod_monto_2012_2021.xlsx"); PP_E2=os.path.join(EX,"expo_paisprod_monto_2022_2025.xlsx")
CP_1=os.path.join(IM,"impo_contipais_monto_2002_2021.xlsx"); CP_2=os.path.join(IM,"impo_contipais_monto_2022_2025.xlsx")
RESUMEN=os.path.join(COMEX,"resumen_intercambiocomercial_2002_2025.xlsx")
TRAF_1=os.path.join(COMEX,"2010_2021_trafico_terrestre_final_v2.xlsx"); TRAF_2=os.path.join(COMEX,"2022_2025_trafico_terrestre_final.xlsx")

NIVEL1_EXP={norm(c["n1_exp"]) for c in CLASIF.values() if c["n1_exp"]}
NIVEL1_IMP={norm(c["n1_imp"]) for c in CLASIF.values() if c["n1_imp"]}

def _read_named(path, sheet=None):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws=wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows=list(ws.iter_rows(values_only=True)); wb.close()
    hdr=None; ycols=[]
    for i,r in enumerate(rows):
        yc=[(j,int(c)) for j,c in enumerate(r) if isinstance(c,int) and 1990<c<2030]
        if len(yc)>=3: hdr=i; ycols=yc; break
    if hdr is None: return [],[]
    namecol=min(j for j,_ in ycols)-1
    out=[]
    for r in rows[hdr+1:]:
        if namecol>=len(r): continue
        nm=r[namecol]
        if nm is None: continue
        s=str(nm).strip()
        if not s or any(t in s for t in ("Fuente","Ir a Listado","Documentos","Declaraciones")): continue
        out.append((s, {y:fnum(r[j]) for j,y in ycols if j<len(r)}))
    return [y for _,y in ycols], out

def _serie(vals): return [round(vals.get(y,0.0),1) for y in MYEARS]

def build_prod_evol(f1,f2,nivset):
    merged={}
    for f in (f1,f2):
        for nm,vals in _read_named(f)[1]:
            merged.setdefault(norm(nm),[nm,{}])[1].update(vals)
    items=[dict(nombre=raw.strip(), serie=_serie(vals)) for k,(raw,vals) in merged.items() if k in nivset]
    items=[i for i in items if any(i["serie"])]
    items.sort(key=lambda i:i["serie"][-1], reverse=True)
    return items

def build_balanza():
    _,rows=_read_named(RESUMEN)
    d={norm(nm):vals for nm,vals in rows}
    def find(sub):
        for k,v in d.items():
            if sub in k: return v
        return {}
    exp=find("EXPORTACION"); cif=find("IMPORTACION (CIF")
    yrs=sorted(y for y in set(list(exp)+list(cif)) if 2002<=y<=2025)
    return dict(years=yrs, exp_fob=[round(exp.get(y,0),1) for y in yrs],
                imp_cif=[round(cif.get(y,0),1) for y in yrs],
                saldo=[round(exp.get(y,0)-cif.get(y,0),1) for y in yrs])

def build_continente_imp():
    CONT={norm(p["continente"]) for p in PAISES.values() if p.get("continente")}
    agg={}
    for f in (CP_1,CP_2):
        for nm,vals in _read_named(f)[1]:
            if norm(nm) in CONT:
                a=agg.setdefault(norm(nm),[nm.strip(),{}]);
                for y,v in vals.items(): a[1][y]=a[1].get(y,0)+v
    items=[dict(nombre=raw, serie=_serie(vals)) for raw,vals in agg.values()]
    items.sort(key=lambda i:i["serie"][-1], reverse=True)
    return items

def parse_paisprod():
    name2cont={norm(p["nombre"]):p.get("continente") for p in PAISES.values()}
    prodset={norm(n) for n,_ in _read_named(PROD_E2)[1]}
    ctot={}; craw={}; comp={}
    for f in (PP_E1,PP_E2):
        cur=None
        for nm,vals in _read_named(f)[1]:
            nk=norm(nm)
            if nk in prodset:
                if cur and nk in NIVEL1_EXP and 2025 in vals:
                    comp.setdefault(cur,{})[nm.strip()]=vals.get(2025,0.0)
            else:
                cur=nk; craw[cur]=nm.strip()
                t=ctot.setdefault(cur,{})
                for y,v in vals.items(): t[y]=t.get(y,0)+v
    # continente export (serie temporal)
    cont={}
    for cnk,vals in ctot.items():
        co=name2cont.get(cnk)
        if not co: continue
        a=cont.setdefault(norm(co),[co,{}])
        for y,v in vals.items(): a[1][y]=a[1].get(y,0)+v
    cont_items=[dict(nombre=raw, serie=_serie(vals)) for raw,vals in cont.values()]
    cont_items.sort(key=lambda i:i["serie"][-1], reverse=True)
    # top paises con composicion 2025
    top=sorted(ctot.items(), key=lambda kv:kv[1].get(2025,0), reverse=True)[:8]
    pp=[]
    for cnk,vals in top:
        comps=sorted(comp.get(cnk,{}).items(), key=lambda kv:kv[1], reverse=True)[:6]
        tot=vals.get(2025,0) or 1
        pp.append(dict(nombre=craw[cnk], fob=round(vals.get(2025,0),1),
                       prod=[dict(nombre=pn, valor=round(pv,1), pct=round(100*pv/tot,1)) for pn,pv in comps]))
    return cont_items, pp

TRAF_YEARS=list(range(2010,2026))
def _read_trafico(path, sheet):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws=wb[sheet]; rows=list(ws.iter_rows(values_only=True)); wb.close()
    hdr=None; ycols=[]
    for i,r in enumerate(rows):
        yc=[(j,int(c)) for j,c in enumerate(r) if isinstance(c,int) and 1990<c<2030]
        if len(yc)>=3: hdr=i; ycols=yc; break
    if hdr is None: return []
    namecol=min(j for j,_ in ycols)-1; flucol=namecol-2
    out=[]; curflu=None
    for r in rows[hdr+1:]:
        if namecol>=len(r): continue
        fl=r[flucol] if 0<=flucol<len(r) else None
        if isinstance(fl,str) and fl.strip(): curflu=fl.strip()
        nm=r[namecol]
        if nm is None: continue
        s=str(nm).strip()
        if not s or any(t in s for t in ("Fuente","Ir a","Total")): continue
        out.append((s, curflu, {y:fnum(r[j]) for j,y in ycols if j<len(r)}))
    return out

def build_trafico():
    data={}
    for path in (TRAF_1,TRAF_2):
        if not os.path.exists(path): continue
        wb=openpyxl.load_workbook(path, read_only=True); sheets=wb.sheetnames; wb.close()
        for metric,kw in (("autos","autom"),("buses","bus"),("camiones","cami"),("carga","carga")):
            shn=next((s for s in sheets if kw in norm(s).lower()), None)
            if not shn: continue
            for nm,flu,vals in _read_trafico(path, shn):
                k=nkey(nm)
                d=data.setdefault(k, dict(raw=nm, autos={}, buses={}, camiones={}, carga={}, cam_ing={}, cam_sal={}))
                for y,v in vals.items(): d[metric][y]=d[metric].get(y,0)+v
                if metric=="camiones":
                    tgt = d["cam_ing"] if (flu and "ING" in norm(flu)) else d["cam_sal"]
                    for y,v in vals.items(): tgt[y]=tgt.get(y,0)+v
    return data
def _serieY(vals): return [round(vals.get(y,0.0),1) for y in TRAF_YEARS]

prod_exp=prod_imp=balanza=cont_imp=cont_exp=paisprod=None; TRAFICO={}
try: prod_exp=build_prod_evol(PROD_E1,PROD_E2,NIVEL1_EXP); prod_imp=build_prod_evol(PROD_I1,PROD_I2,NIVEL1_IMP)
except Exception as e: print("  [prod] ", e)
try: balanza=build_balanza()
except Exception as e: print("  [balanza] ", e)
try: cont_imp=build_continente_imp()
except Exception as e: print("  [cont_imp] ", e)
try: cont_exp,paisprod=parse_paisprod()
except Exception as e: print("  [paisprod] ", e)
try: TRAFICO=build_trafico()
except Exception as e: print("  [trafico] ", e)

# adjuntar trafico a cada punto terrestre (match por nombre normalizado)
traf_n=0
for r in registros:
    if r["grupo"]!="terrestre": continue
    pk=nkey(r["nombre"]); match=None
    for tk in TRAFICO:
        if pk==tk: match=tk; break
    if not match:
        for tk in TRAFICO:
            if len(pk)>=5 and (pk in tk or tk in pk): match=tk; break
    if match:
        t=TRAFICO[match]
        r["trafico"]=dict(years=TRAF_YEARS, raw=t["raw"],
            camiones_ing=_serieY(t["cam_ing"]), camiones_sal=_serieY(t["cam_sal"]),
            autos=_serieY(t["autos"]), buses=_serieY(t["buses"]), carga=_serieY(t["carga"]))
        traf_n+=1

MACRO=dict(
    prod=dict(years=MYEARS, exp=(prod_exp or [])[:8], imp=(prod_imp or [])[:8]),
    continente=dict(years=MYEARS, imp=cont_imp or [], exp=cont_exp or []),
    paisprod=paisprod or [],
    balanza=balanza or {},
)
print(f"  prod_exp={len(prod_exp or [])} cont_imp={len(cont_imp or [])} paisprod={len(paisprod or [])} balanza_anios={len((balanza or {}).get('years',[]))} trafico_puntos={traf_n}", flush=True)

with open(os.path.join(DATA,"regiones_perfil.json"),"w",encoding="utf-8") as f:
    json.dump(regiones_perfil, f, ensure_ascii=False, indent=1)

bundle = dict(meta=meta, nacional=nacional, puntos=registros, paises=paises_comercio,
              world=world, regiones=chreg, region_nombres=REGIONES, macro=MACRO,
              regiones_perfil=regiones_perfil)
with open(os.path.join(HERE,"data_bundle.js"),"w",encoding="utf-8") as f:
    f.write("window.DATA = ")
    json.dump(bundle, f, ensure_ascii=False)
    f.write(";\n")

# versiona la referencia al bundle en los HTML (cache-busting: evita HTML nuevo + datos viejos)
# y estampa la insignia de build visible en el pie (diagnostico de version en un vistazo)
import datetime
_bhash = hashlib.md5(open(os.path.join(HERE,"data_bundle.js"),"rb").read()).hexdigest()[:10]
_build = datetime.datetime.now().strftime("%Y-%m-%d %H:%M") + " · " + _bhash[:6]
for _page in ("index.html", "mapa_puertos.html"):
    _pp = os.path.join(HERE, _page)
    if os.path.exists(_pp):
        _t = open(_pp, encoding="utf-8").read()
        _t2 = re.sub(r'data_bundle\.js(\?v=[a-f0-9]+)?', f'data_bundle.js?v={_bhash}', _t)
        _t2 = re.sub(r'build (BUILD_TAG|[\d\-]+ [\d:]+ · [a-f0-9]+)', f'build {_build}', _t2)
        if _t2 != _t:
            open(_pp, "w", encoding="utf-8").write(_t2)
print(f"  bundle v={_bhash} · build '{_build}' estampados en HTML", flush=True)

# ---------------------------------------------------------------- reporte
print("\n================ CUADRE DE CONTROL ================")
print(f"Exportacion nacional (todas ops):  US$ {nat_fob_all:,.0f}")
print(f"Suma FOB puntos chilenos:          US$ {sum_fob:,.0f}  ({meta['cuadre']['exp_cobertura_pct']}%)")
print(f"Importacion nacional (todas ops):  US$ {nat_cif_all:,.0f}")
print(f"Suma CIF puntos chilenos:          US$ {sum_cif:,.0f}  ({meta['cuadre']['imp_cobertura_pct']}%)")
print(f"\nPuntos con comercio: {len(registros)}")
print(f"Puntos SIN coordenada ({len(sin_coord)}):")
for s in sin_coord:
    print(f"   - [{s['cod']}] {s['nombre']} ({s['tipo']})  total US$ {s['total']:,.0f}")
print("\nArchivos escritos en:", DATA)
print("Bundle:", os.path.join(HERE,"data_bundle.js"))
